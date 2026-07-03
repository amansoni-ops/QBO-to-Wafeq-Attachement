"""
app.py — QBO to Wafeq Migration Tool
=====================================
Credentials are loaded from .env file via config.py.
Never hardcode secrets here.

Setup:
  1. Copy .env.example → .env
  2. Fill in your values
  3. python start.py
"""

import json
import base64
import secrets
import logging
import threading
from pathlib import Path
from datetime import datetime, timedelta, timezone
from urllib.parse import urlencode

import requests
from flask import Flask, request, jsonify, send_from_directory, Response, session
from flask_cors import CORS

# ── Logging — writes to app.log + console ─────────────────────────────────────
_log_file = Path("app.log")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s — %(message)s",
    datefmt="%Y-%m-%d %H:%M:%S",
    handlers=[
        logging.FileHandler(_log_file, encoding="utf-8"),
        logging.StreamHandler(),
    ]
)
log = logging.getLogger("qbo_wafeq")

from config import (
    CLIENT_ID, CLIENT_SECRET, ENVIRONMENT,
    REDIRECT_URI, REDIRECT_PATH, FLASK_PORT,
    INTUIT_AUTH_URL, INTUIT_TOKEN_URL, QB_BASE_URLS,
)

# ---------------------------------------------------------------------------
# App-level constants (not credentials)
# ---------------------------------------------------------------------------
TOKENS_DIR = Path("qb_profiles")
DB_FILE    = Path("companies.json")

# state → pending company info
_pending_oauth: dict = {}

app = Flask(__name__, static_folder="frontend_dist", static_url_path="")

@app.after_request
def _log_request(resp):
    # Log every API call to app.log
    if request.path.startswith("/api"):
        user = session.get("user", "anonymous")
        log.info(f'{user} {request.method} {request.path} → {resp.status_code}')
    return resp

app.secret_key = __import__('os').environ.get('FLASK_SECRET_KEY', 'dev-secret-key-change-in-prod')
CORS(app, supports_credentials=True)

TOKENS_DIR.mkdir(exist_ok=True)


# ---------------------------------------------------------------------------
# Companies DB
# ---------------------------------------------------------------------------

def load_db() -> dict:
    if not DB_FILE.exists():
        return {}
    return json.loads(DB_FILE.read_text())

def save_db(data: dict):
    DB_FILE.write_text(json.dumps(data, indent=2))

def upsert_company(realm_id: str, data: dict):
    db = load_db()
    db[realm_id] = {**db.get(realm_id, {}), **data}
    save_db(db)


# ---------------------------------------------------------------------------
# Token helpers
# ---------------------------------------------------------------------------

def _basic_header():
    return "Basic " + base64.b64encode(
        f"{CLIENT_ID}:{CLIENT_SECRET}".encode()
    ).decode()

def token_path(realm_id: str) -> Path:
    return TOKENS_DIR / f"qb_tokens_{realm_id}.json"

def load_tokens(realm_id: str) -> dict | None:
    p = token_path(realm_id)
    if not p.exists():
        return None
    return json.loads(p.read_text())

def save_tokens(realm_id: str, data: dict):
    token_path(realm_id).write_text(json.dumps(data, indent=2))

def is_token_valid(tokens: dict, buffer_sec: int = 120) -> bool:
    try:
        exp = datetime.fromisoformat(tokens["access_token_expires_at"])
        if exp.tzinfo is None:
            exp = exp.replace(tzinfo=timezone.utc)
        return datetime.now(timezone.utc) < (exp - timedelta(seconds=buffer_sec))
    except:
        return False

def do_refresh(realm_id: str, stored: dict) -> dict:
    resp = requests.post(
        INTUIT_TOKEN_URL,
        headers={
            "Authorization": _basic_header(),
            "Content-Type":  "application/x-www-form-urlencoded",
            "Accept":        "application/json",
        },
        data={
            "grant_type":    "refresh_token",
            "refresh_token": stored["refresh_token"],
        },
        timeout=15,
    )
    if resp.status_code != 200:
        raise RuntimeError(f"Refresh failed ({resp.status_code}): {resp.text}")
    t = resp.json()
    updated = {
        **stored,
        "access_token":            t["access_token"],
        "refresh_token":           t["refresh_token"],
        "access_token_expires_at": (
            datetime.now(timezone.utc) + timedelta(seconds=t.get("expires_in", 3600))
        ).isoformat(),
        "saved_at": datetime.now(timezone.utc).isoformat(),
    }
    save_tokens(realm_id, updated)
    return updated

def get_access_token(realm_id: str) -> str:
    tokens = load_tokens(realm_id)
    if not tokens:
        raise RuntimeError(f"No tokens for {realm_id}. Connect first.")
    if not is_token_valid(tokens):
        tokens = do_refresh(realm_id, tokens)
    return tokens["access_token"]


# ---------------------------------------------------------------------------
# Fetch company info from QB API
# ---------------------------------------------------------------------------

def fetch_company_info(access_token: str, realm_id: str) -> dict:
    base = QB_BASE_URLS[ENVIRONMENT]
    url  = f"{base}/v3/company/{realm_id}/companyinfo/{realm_id}"
    resp = requests.get(
        url,
        headers={
            "Authorization": f"Bearer {access_token}",
            "Accept":        "application/json",
        },
        params={"minorversion": "65"},
        timeout=15,
    )
    resp.raise_for_status()
    return resp.json().get("CompanyInfo", {})


# ---------------------------------------------------------------------------
# Serve frontend
# ---------------------------------------------------------------------------

@app.route("/")
def index():
    return send_from_directory("frontend_dist", "index.html")


# ---------------------------------------------------------------------------
# API — Get all connected companies (for dropdown)
# ---------------------------------------------------------------------------

@app.route("/api/companies", methods=["GET"])
def api_list_companies():
    from core.storage import get_stats
    db     = load_db()
    result = []

    for realm_id, info in db.items():
        tokens      = load_tokens(realm_id)
        token_valid = is_token_valid(tokens) if tokens else False
        result.append({
            "realm_id":    realm_id,
            "name":        info.get("name", f"Company {realm_id}"),
            "country":     info.get("country", "??"),
            "environment": info.get("environment", ENVIRONMENT),
            "connected":   bool(tokens),
            "token_valid": token_valid,
            "stats":       get_stats(realm_id),
            "wafeq_api_key":  info.get("wafeq_api_key", ""),
            "wafeq_key_name": info.get("wafeq_key_name", ""),
            "wafeq_keys":     info.get("wafeq_keys", []),
        })

    result.sort(key=lambda x: (not x["connected"], x["name"]))
    return jsonify(result)


# ---------------------------------------------------------------------------
# API — Start OAuth flow
# ---------------------------------------------------------------------------

@app.route("/api/qbo/auth-url", methods=["GET"])
def api_get_auth_url():
    state = secrets.token_urlsafe(32)
    _pending_oauth[state] = {"started_at": datetime.now().isoformat()}

    auth_url = INTUIT_AUTH_URL + "?" + urlencode({
        "client_id":     CLIENT_ID,
        "response_type": "code",
        "scope":         "com.intuit.quickbooks.accounting",
        "redirect_uri":  REDIRECT_URI,
        "state":         state,
    })
    return jsonify({"auth_url": auth_url})


# ---------------------------------------------------------------------------
# API — OAuth callback (QB redirects here after login)
# ---------------------------------------------------------------------------

@app.route("/callback", methods=["GET"])
def api_callback():
    auth_code = request.args.get("code")
    realm_id  = request.args.get("realmId")
    state     = request.args.get("state")
    error     = request.args.get("error")

    if error:
        return _callback_page("error", f"Authorization failed: {error}")

    if not auth_code or not realm_id:
        return _callback_page("error", "Missing parameters from QB")

    try:
        resp = requests.post(
            INTUIT_TOKEN_URL,
            headers={
                "Authorization": _basic_header(),
                "Content-Type":  "application/x-www-form-urlencoded",
                "Accept":        "application/json",
            },
            data={
                "grant_type":   "authorization_code",
                "code":         auth_code,
                "redirect_uri": REDIRECT_URI,
            },
            timeout=15,
        )
        if resp.status_code != 200:
            raise RuntimeError(f"Token exchange failed ({resp.status_code}): {resp.text}")

        t = resp.json()
        tokens = {
            "access_token":               t["access_token"],
            "refresh_token":              t["refresh_token"],
            "realm_id":                   realm_id,
            "environment":                ENVIRONMENT,
            "access_token_expires_at":    (
                datetime.now(timezone.utc) + timedelta(seconds=t.get("expires_in", 3600))
            ).isoformat(),
            "x_refresh_token_expires_in": t.get("x_refresh_token_expires_in"),
            "saved_at":                   datetime.now(timezone.utc).isoformat(),
        }
        save_tokens(realm_id, tokens)

        try:
            info = fetch_company_info(t["access_token"], realm_id)
            company_name    = info.get("CompanyName", f"Company {realm_id}")
            company_country = info.get("Country", "??")
        except:
            company_name    = f"Company {realm_id}"
            company_country = "??"

        upsert_company(realm_id, {
            "name":        company_name,
            "country":     company_country,
            "environment": ENVIRONMENT,
            "realm_id":    realm_id,
            "connected":   True,
            "stats":       {},
        })

        _pending_oauth.pop(state, None)
        return _callback_page("success", company_name, realm_id)

    except Exception as e:
        return _callback_page("error", str(e))


def _callback_page(status: str, message: str, realm_id: str = "") -> str:
    if status == "success":
        icon  = "✓"
        color = "#2dd98f"
        title = "Connected!"
        body  = f"<b>{message}</b> has been connected successfully."
    else:
        icon  = "✕"
        color = "#ff5c5c"
        title = "Connection failed"
        body  = message

    return f"""
    <html>
    <head>
      <link href="https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500&display=swap" rel="stylesheet">
      <style>
        * {{ margin:0;padding:0;box-sizing:border-box }}
        body {{ font-family:'DM Sans',sans-serif;background:#0d0f14;color:#e8eaf0;
               display:flex;align-items:center;justify-content:center;min-height:100vh; }}
        .card {{ text-align:center;padding:48px 40px;border:1px solid #262d3d;
                border-radius:16px;background:#151820;max-width:380px;width:90%; }}
        .icon {{ font-size:44px;margin-bottom:16px;color:{color} }}
        h2 {{ font-size:20px;font-weight:600;color:{color};margin-bottom:10px }}
        p {{ color:#6b748a;font-size:13px;line-height:1.6 }}
        .note {{ margin-top:20px;font-size:11px;font-family:monospace;color:#4f7fff }}
      </style>
      <script>
        setTimeout(() => {{
          if (window.opener) {{
            window.opener.postMessage({{
              type: 'QB_AUTH_{status.upper()}',
              realm_id: '{realm_id}',
              name: '{message}'
            }}, '*');
          }}
          window.close();
        }}, 2000);
      </script>
    </head>
    <body>
      <div class="card">
        <div class="icon">{icon}</div>
        <h2>{title}</h2>
        <p>{body}</p>
        <div class="note">closing automatically...</div>
      </div>
    </body>
    </html>
    """


# ---------------------------------------------------------------------------
# API — Test connection
# ---------------------------------------------------------------------------

@app.route("/api/qbo/test/<realm_id>", methods=["GET"])
def api_test(realm_id: str):
    try:
        access_token = get_access_token(realm_id)
        info         = fetch_company_info(access_token, realm_id)
        name         = info.get("CompanyName", realm_id)
        country      = info.get("Country", "??")
        upsert_company(realm_id, {"name": name, "country": country})
        return jsonify({"ok": True, "name": name, "country": country})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# API — Delete company
# ---------------------------------------------------------------------------

@app.route("/api/companies/<realm_id>", methods=["DELETE"])
def api_delete_company(realm_id: str):
    db = load_db()
    if realm_id not in db:
        return jsonify({"error": "Not found"}), 404
    del db[realm_id]
    save_db(db)
    p = token_path(realm_id)
    if p.exists():
        p.unlink()
    return jsonify({"ok": True})


# ---------------------------------------------------------------------------
# API — Fetch bills + attachments (SSE live progress stream)
# ---------------------------------------------------------------------------

@app.route("/api/fetch/<realm_id>", methods=["GET"])
def api_fetch(realm_id: str):
    """
    Fetch QB transactions + attachments via SSE stream.
    Query params:
      ?limit=N             → fetch only first N transactions (for testing)
      ?clear=1             → delete existing index before fetching clean
      ?date_from=YYYY-MM-DD → filter transactions from this date
      ?date_to=YYYY-MM-DD   → filter transactions up to this date
      ?types=Bill,Invoice   → comma-separated entity types to fetch
                              (blank = all supported types)
    """
    import queue
    from core.fetcher import fetch_and_store, SUPPORTED_ENTITY_TYPES
    from core.storage import clear_index

    limit      = request.args.get("limit", "").strip()
    clear      = request.args.get("clear", "0") == "1"
    date_from  = request.args.get("date_from", "").strip() or None
    date_to    = request.args.get("date_to", "").strip()   or None
    types_raw  = request.args.get("types", "").strip()
    bill_limit = int(limit) if limit.isdigit() and int(limit) > 0 else None

    # Parse entity type filter — validate against known types
    entity_types = None
    if types_raw:
        requested = {t.strip() for t in types_raw.split(",") if t.strip()}
        entity_types = requested & SUPPORTED_ENTITY_TYPES  # only valid ones
        if not entity_types:
            entity_types = None  # invalid filter → fetch all

    if clear:
        clear_index(realm_id)

    q = queue.Queue()

    def progress(msg, t="info"):
        q.put({"msg": msg, "type": t})

    def run():
        try:
            fetch_and_store(realm_id, progress_cb=progress, bill_limit=bill_limit,
                            date_from=date_from, date_to=date_to,
                            entity_types=entity_types)
        except Exception as e:
            q.put({"msg": f"ERROR: {e}", "type": "err"})
        finally:
            q.put(None)

    threading.Thread(target=run, daemon=True).start()

    def stream():
        while True:
            item = q.get()
            if item is None:
                yield "data: __DONE__\n\n"
                break
            yield f"data: {json.dumps(item)}\n\n"

    return Response(
        stream(),
        mimetype="text/event-stream",
        headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"},
    )


# ---------------------------------------------------------------------------
# API — Get bill+attachment index (already fetched)
# ---------------------------------------------------------------------------

@app.route("/api/index/<realm_id>", methods=["GET"])
def api_index(realm_id: str):
    from core.storage import load_index
    return jsonify(load_index(realm_id))


# ---------------------------------------------------------------------------
# Auth routes — must be before if __name__ block
# ---------------------------------------------------------------------------

from core.auth_users import (
    authenticate, get_user, list_users, create_user,
    change_password, admin_reset_password, delete_user,
    log_action, get_audit_log
)

@app.route('/api/auth/login', methods=['POST'])
def api_login():
    body     = request.json or {}
    username = body.get('username', '').strip()
    password = body.get('password', '').strip()
    user     = authenticate(username, password)
    if user:
        session['user'] = username
        session['role'] = user['role']
        log_action(username, 'login', {'ip': request.remote_addr})
        return jsonify({'ok': True, 'user': username, 'role': user['role']})
    return jsonify({'error': 'Invalid username or password'}), 401

@app.route('/api/auth/logout', methods=['POST'])
def api_logout():
    user = session.get('user')
    if user:
        log_action(user, 'logout')
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/auth/me', methods=['GET'])
def api_me():
    username = session.get('user')
    if not username:
        return jsonify({'error': 'Not logged in'}), 401
    u = get_user(username)
    return jsonify({'user': username, 'role': session.get('role', 'user'), 'info': u})

@app.route('/api/auth/change-password', methods=['POST'])
def api_change_password():
    username = session.get('user')
    if not username:
        return jsonify({'error': 'Not logged in'}), 401
    body        = request.json or {}
    old_pw      = body.get('old_password', '').strip()
    new_pw      = body.get('new_password', '').strip()
    ok, msg     = change_password(username, old_pw, new_pw)
    if ok:
        log_action(username, 'change_password')
        return jsonify({'ok': True, 'message': msg})
    return jsonify({'error': msg}), 400

@app.route('/api/admin/users', methods=['GET'])
def api_list_users():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    return jsonify(list_users())

@app.route('/api/admin/users', methods=['POST'])
def api_create_user():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    body     = request.json or {}
    username = body.get('username', '').strip()
    password = body.get('password', '').strip()
    role     = body.get('role', 'user').strip()
    ok, msg  = create_user(username, password, role)
    if ok:
        log_action(session['user'], 'create_user', {'target': username, 'role': role})
        return jsonify({'ok': True, 'message': msg})
    return jsonify({'error': msg}), 400

@app.route('/api/admin/users/<target>/reset-password', methods=['POST'])
def api_reset_password(target: str):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    body     = request.json or {}
    new_pw   = body.get('new_password', '').strip()
    ok, msg  = admin_reset_password(session['user'], target, new_pw)
    if ok:
        log_action(session['user'], 'reset_password', {'target': target})
        return jsonify({'ok': True, 'message': msg})
    return jsonify({'error': msg}), 400

@app.route('/api/admin/users/<target>', methods=['DELETE'])
def api_delete_user(target: str):
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    ok, msg = delete_user(session['user'], target)
    if ok:
        log_action(session['user'], 'delete_user', {'target': target})
        return jsonify({'ok': True, 'message': msg})
    return jsonify({'error': msg}), 400

@app.route('/api/admin/audit', methods=['GET'])
def api_audit_log():
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    limit    = int(request.args.get('limit', 500))
    username = request.args.get('user')
    return jsonify(get_audit_log(limit=limit, username=username))

@app.route('/api/admin/report', methods=['GET'])
def api_admin_report():
    """Full report: users, audit log, company stats, app.log tail — admin only."""
    if session.get('role') != 'admin':
        return jsonify({'error': 'Admin only'}), 403
    from core.storage import load_companies
    db = load_companies()

    # Read last 200 lines of app.log
    app_log_lines = []
    try:
        log_path = Path("app.log")
        if log_path.exists():
            with open(log_path, "r", encoding="utf-8") as f:
                app_log_lines = f.readlines()[-200:]
    except Exception:
        pass

    return jsonify({
        'users':     list_users(),
        'audit':     get_audit_log(limit=1000),
        'app_log':   [l.rstrip() for l in app_log_lines],
        'companies': [
            {
                'realm_id':    rid,
                'name':        info.get('name'),
                'environment': info.get('environment'),
                'stats':       info.get('stats', {}),
            }
            for rid, info in db.items()
        ],
    })


# ---------------------------------------------------------------------------
# Run
# ---------------------------------------------------------------------------


# ---------------------------------------------------------------------------
# API — Wafeq: save API key
# ---------------------------------------------------------------------------

@app.route("/api/wafeq/key", methods=["POST"])
def api_save_wafeq_key():
    """
    Add a new Wafeq API key to a company's key list.
    If set_active=true, makes it the active key.
    Supports multiple named keys per company.
    """
    body      = request.json
    realm_id  = body.get("realm_id", "").strip()
    api_key   = body.get("api_key", "").strip()
    key_name  = body.get("key_name", "").strip() or "Key"
    set_active= body.get("set_active", True)

    if not realm_id or not api_key:
        return jsonify({"error": "realm_id and api_key required"}), 400

    db   = load_db()
    info = db.get(realm_id, {})

    # Get existing keys list
    keys = info.get("wafeq_keys", [])

    # Avoid duplicate keys
    existing = next((k for k in keys if k["key"] == api_key), None)
    if existing:
        existing["name"] = key_name
    else:
        import uuid
        keys.append({"id": str(uuid.uuid4())[:8], "name": key_name, "key": api_key})

    updates = {"wafeq_keys": keys}

    # Set as active key if requested
    if set_active:
        updates["wafeq_api_key"]  = api_key
        updates["wafeq_key_name"] = key_name

    upsert_company(realm_id, updates)
    return jsonify({"ok": True, "total_keys": len(keys)})


@app.route("/api/wafeq/key/activate", methods=["POST"])
def api_activate_wafeq_key():
    """Set a specific saved key as the active key for migration."""
    body     = request.json
    realm_id = body.get("realm_id", "").strip()
    key_id   = body.get("key_id", "").strip()
    if not realm_id or not key_id:
        return jsonify({"error": "realm_id and key_id required"}), 400
    db   = load_db()
    info = db.get(realm_id, {})
    keys = info.get("wafeq_keys", [])
    key  = next((k for k in keys if k["id"] == key_id), None)
    if not key:
        return jsonify({"error": "Key not found"}), 404
    upsert_company(realm_id, {
        "wafeq_api_key":  key["key"],
        "wafeq_key_name": key["name"],
    })
    return jsonify({"ok": True, "active_name": key["name"]})


@app.route("/api/wafeq/key/delete", methods=["POST"])
def api_delete_wafeq_key():
    """Remove a saved key from a company's key list."""
    body     = request.json
    realm_id = body.get("realm_id", "").strip()
    key_id   = body.get("key_id", "").strip()
    if not realm_id or not key_id:
        return jsonify({"error": "realm_id and key_id required"}), 400
    db   = load_db()
    info = db.get(realm_id, {})
    keys = [k for k in info.get("wafeq_keys", []) if k["id"] != key_id]
    updates = {"wafeq_keys": keys}
    # If active key was deleted, clear it
    deleted = next((k for k in info.get("wafeq_keys", []) if k["id"] == key_id), None)
    if deleted and info.get("wafeq_api_key") == deleted["key"]:
        # Set first remaining key as active, or clear
        if keys:
            updates["wafeq_api_key"]  = keys[0]["key"]
            updates["wafeq_key_name"] = keys[0]["name"]
        else:
            updates["wafeq_api_key"]  = ""
            updates["wafeq_key_name"] = ""
    upsert_company(realm_id, updates)
    return jsonify({"ok": True, "remaining": len(keys)})


@app.route("/api/wafeq/test/<realm_id>", methods=["GET"])
def api_test_wafeq(realm_id: str):
    """Test Wafeq API key by fetching one bill."""
    db  = load_db()
    key = db.get(realm_id, {}).get("wafeq_api_key", "")
    if not key:
        return jsonify({"error": "No Wafeq API key set for this company."}), 400
    try:
        import requests as req
        r = req.get(
            "https://api.wafeq.com/v1/bills/",
            headers={"Authorization": f"Api-Key {key}", "Accept": "application/json"},
            params={"limit": 1},
            timeout=10,
        )
        if r.status_code == 401:
            return jsonify({"error": "Invalid API key — Wafeq returned 401."}), 401
        r.raise_for_status()
        count = r.json().get("count", 0)
        return jsonify({"ok": True, "bill_count": count})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


# ---------------------------------------------------------------------------
# API — Wafeq: fetch bills + match
# ---------------------------------------------------------------------------

@app.route("/api/match/<realm_id>", methods=["GET"])
def api_match(realm_id: str):
    """Fetch Wafeq bills and match with QB bills via SSE stream."""
    import queue
    from core.matcher import run_matching

    db  = load_db()
    key = db.get(realm_id, {}).get("wafeq_api_key", "")
    if not key:
        return jsonify({"error": "No Wafeq API key. Set it in Settings first."}), 400

    expense_target = request.args.get("expense_target", "expense").strip()

    q = queue.Queue()

    def progress(msg, t="info"):
        q.put({"msg": msg, "type": t})

    def run():
        try:
            run_matching(realm_id, key, progress_cb=progress, expense_target=expense_target)
        except Exception as e:
            q.put({"msg": f"ERROR: {e}", "type": "err"})
        finally:
            q.put(None)

    threading.Thread(target=run, daemon=True).start()

    def stream():
        while True:
            item = q.get()
            if item is None:
                yield "data: __DONE__\n\n"
                break
            yield f"data: {json.dumps(item)}\n\n"

    return Response(stream(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# API — Upload matched attachments to Wafeq
# ---------------------------------------------------------------------------

@app.route("/api/upload/<realm_id>", methods=["GET"])
def api_upload(realm_id: str):
    """
    Upload matched attachments to Wafeq via SSE stream.
    Query params:
      ?retry_failed=1       → retry only previously failed files
      ?bill_id={qb_bill_id} → upload only one specific bill (single bill retry)
    """
    import queue
    from core.uploader import run_upload

    db  = load_db()
    key = db.get(realm_id, {}).get("wafeq_api_key", "")
    if not key:
        return jsonify({"error": "No Wafeq API key. Set it in the sidebar first."}), 400

    retry_failed = request.args.get("retry_failed", "0") == "1"
    single_bill  = request.args.get("bill_id", "").strip() or None

    q = queue.Queue()

    def progress(msg, t="info"):
        q.put({"msg": msg, "type": t})

    def run():
        try:
            run_upload(realm_id, key, progress_cb=progress,
                       retry_failed=retry_failed, single_bill_id=single_bill)
        except Exception as e:
            q.put({"msg": f"ERROR: {e}", "type": "err"})
        finally:
            q.put(None)

    threading.Thread(target=run, daemon=True).start()

    def stream():
        while True:
            item = q.get()
            if item is None:
                yield "data: __DONE__\n\n"
                break
            yield f"data: {json.dumps(item)}\n\n"

    return Response(stream(), mimetype="text/event-stream",
                    headers={"Cache-Control": "no-cache", "X-Accel-Buffering": "no"})


# ---------------------------------------------------------------------------
# API — Manual match override
# ---------------------------------------------------------------------------

@app.route("/api/manual-match", methods=["POST"])
def api_manual_match():
    """
    Manually assign a Wafeq record to any QB transaction type.
    Looks up the Wafeq record to return its doc/ref number for confirmation.
    """
    from core.storage import load_index, save_index
    import requests as req

    body          = request.json
    realm_id      = body.get("realm_id")
    qb_bill_id    = body.get("qb_bill_id")
    wafeq_bill_id = body.get("wafeq_bill_id")
    wafeq_type    = body.get("wafeq_type", "bill").lower()

    if not all([realm_id, qb_bill_id, wafeq_bill_id]):
        return jsonify({"error": "realm_id, qb_bill_id, wafeq_bill_id required"}), 400

    idx  = load_index(realm_id)
    bill = idx["bills"].get(qb_bill_id)
    if not bill:
        return jsonify({"error": "QB transaction not found in index"}), 404

    # Endpoint map for looking up the Wafeq record doc number
    endpoint_map = {
        "bill":    "bills",
        "invoice": "invoices",
        "journal": "journal-entries",
    }
    endpoint = endpoint_map.get(wafeq_type, "bills")

    # Try to fetch the Wafeq record to get its doc/ref number
    wafeq_doc_number = None
    db  = load_db()
    key = db.get(realm_id, {}).get("wafeq_api_key", "")
    if key:
        try:
            from config import WAFEQ_BASE_URL
            r = req.get(
                f"{WAFEQ_BASE_URL}/{endpoint}/{wafeq_bill_id}/",
                headers={"Authorization": f"Api-Key {key}", "Accept": "application/json"},
                timeout=8,
            )
            if r.ok:
                d = r.json()
                wafeq_doc_number = (
                    d.get("bill_number") or d.get("invoice_number") or
                    d.get("number") or d.get("reference") or
                    d.get("journal_number") or None
                )
        except Exception:
            pass

    bill["wafeq_bill_id"]  = wafeq_bill_id
    bill["wafeq_type"]     = wafeq_type
    bill["match_status"]   = "manual"
    bill["match_note"]     = f"Manually assigned to {wafeq_type} {wafeq_bill_id}"
    if wafeq_doc_number:
        bill["match_note"] += f" (ref: {wafeq_doc_number})"

    save_index(realm_id, idx)
    return jsonify({
        "ok": True,
        "wafeq_doc_number": wafeq_doc_number,
        "wafeq_type": wafeq_type,
    })


# ---------------------------------------------------------------------------
# API — Export migration report as Excel (2 sheets: Bills + Attachments)
# ---------------------------------------------------------------------------

@app.route("/api/report/<realm_id>", methods=["GET"])
def api_report(realm_id: str):
    """
    Generate migration report for a company.
    Returns JSON summary + per-transaction data + type breakdown.
    ?format=xlsx → returns downloadable Excel report.
    """
    from core.storage import load_index
    import io

    idx   = load_index(realm_id)
    bills = idx.get("bills", {})

    # Build summary
    total      = len(bills)
    downloaded = sum(1 for b in bills.values() for a in b.get("attachments",[]) if a.get("download_status")=="success")
    matched    = sum(1 for b in bills.values() if b.get("match_status")=="matched")
    no_match   = sum(1 for b in bills.values() if b.get("match_status")=="no_match")
    uploaded   = sum(1 for b in bills.values() for a in b.get("attachments",[]) if a.get("upload_status")=="success")
    failed     = sum(1 for b in bills.values() for a in b.get("attachments",[]) if a.get("upload_status")=="failed")
    total_files= sum(len(b.get("attachments",[])) for b in bills.values())

    # Type breakdown
    from collections import defaultdict
    from core.matcher import QBO_TO_WAFEQ_TYPE
    type_data = defaultdict(lambda: {"count":0,"matched":0,"uploaded":0,"failed":0,"wafeq_type":""})
    for b in bills.values():
        qtype = b.get("qbo_type","Bill")
        wtype = QBO_TO_WAFEQ_TYPE.get(qtype.lower().replace(" ",""), "bill")
        type_data[qtype]["count"] += 1
        type_data[qtype]["wafeq_type"] = wtype
        if b.get("match_status") == "matched": type_data[qtype]["matched"] += 1
        for a in b.get("attachments",[]):
            if a.get("upload_status")=="success": type_data[qtype]["uploaded"] += 1
            if a.get("upload_status")=="failed":  type_data[qtype]["failed"]  += 1

    type_breakdown = [{"qbo_type":k,**v} for k,v in type_data.items()]
    type_breakdown.sort(key=lambda x: -x["count"])

    summary = {
        "total":       total,
        "total_files": total_files,
        "downloaded":  downloaded,
        "matched":     matched,
        "no_match":    no_match,
        "uploaded":    uploaded,
        "failed":      failed,
        "fetched_at":  idx.get("fetched_at"),
    }

    bill_list = list(bills.values())

    # Return Excel if requested
    fmt = request.args.get("format", "json")
    if fmt == "xlsx":
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        wb = Workbook()

        # Summary sheet
        ws1 = wb.active
        ws1.title = "Summary"
        hdr = Font(name="Arial", bold=True, color="FFFFFF")
        hfill = PatternFill("solid", fgColor="1F3864")
        ctr = Alignment(horizontal="center", vertical="center")

        ws1.merge_cells("A1:D1")
        ws1["A1"] = "Migration Summary Report"
        ws1["A1"].font = Font(name="Arial", bold=True, size=14, color="FFFFFF")
        ws1["A1"].fill = hfill
        ws1["A1"].alignment = ctr
        ws1.row_dimensions[1].height = 30

        for r, (k,v) in enumerate([
            ("Company Realm ID", realm_id),
            ("Report generated", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Total Transactions", total),
            ("Files Downloaded", downloaded),
            ("Matched", matched),
            ("No Match", no_match),
            ("Uploaded", uploaded),
            ("Failed", failed),
        ], start=3):
            ws1.cell(row=r, column=1, value=k).font = Font(name="Arial", bold=True, size=10)
            ws1.cell(row=r, column=2, value=v).font = Font(name="Arial", size=10)
        ws1.column_dimensions["A"].width = 28
        ws1.column_dimensions["B"].width = 35

        # Type breakdown sheet
        ws2 = wb.create_sheet("By Type")
        hdrs = ["QBO Type","Wafeq Target","Count","Matched","Uploaded","Failed","Status"]
        for ci, h in enumerate(hdrs, 1):
            c = ws2.cell(row=1, column=ci, value=h)
            c.font = hdr; c.fill = hfill; c.alignment = ctr
        for ri, row in enumerate(type_breakdown, 2):
            vals = [row["qbo_type"], row["wafeq_type"], row["count"], row["matched"],
                    row["uploaded"], row["failed"],
                    "Complete" if row["count"]==row["uploaded"] else "Partial" if row["uploaded"]>0 else "Pending"]
            for ci, v in enumerate(vals, 1):
                ws2.cell(row=ri, column=ci, value=v).font = Font(name="Arial", size=10)
        for col in ["A","B","C","D","E","F","G"]:
            ws2.column_dimensions[col].width = 20

        # Transactions sheet
        ws3 = wb.create_sheet("Transactions")
        th = ["Doc #","QBO Type","Vendor/Contact","Date","Amount","Currency",
              "Match Status","Wafeq Type","Wafeq ID","Total Files","Uploaded","Failed","Notes"]
        for ci, h in enumerate(th, 1):
            c = ws3.cell(row=1, column=ci, value=h)
            c.font = hdr; c.fill = hfill; c.alignment = ctr
        for ri, b in enumerate(bill_list, 2):
            atts = b.get("attachments",[])
            up   = sum(1 for a in atts if a.get("upload_status")=="success")
            fa   = sum(1 for a in atts if a.get("upload_status")=="failed")
            row  = [
                b.get("doc_number",""), b.get("qbo_type","Bill"),
                b.get("vendor_name",""), b.get("txn_date",""),
                b.get("total_amt",0), b.get("currency",""),
                b.get("match_status","pending"), b.get("wafeq_type",""),
                b.get("wafeq_bill_id",""), len(atts), up, fa,
                b.get("match_note",""),
            ]
            for ci, v in enumerate(row, 1):
                ws3.cell(row=ri, column=ci, value=v).font = Font(name="Arial", size=10)
        col_widths = [14,20,28,12,12,10,16,14,26,12,10,8,40]
        for ci, w in enumerate(col_widths, 1):
            ws3.column_dimensions[chr(64+ci)].width = w

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=migration_report_{realm_id}.xlsx"}
        )

    return jsonify({
        "summary":        summary,
        "type_breakdown": type_breakdown,
        "bills":          bill_list,
    })


# QBO type → human label + contact-column label (for Excel export)
QBO_TYPE_LABELS = {
    "bill":              ("Bill",              "Vendor Name"),
    "invoice":           ("Invoice",           "Customer Name"),
    "creditmemo":        ("Credit Note",       "Customer Name"),
    "vendorcredit":      ("Debit Note",        "Vendor Name"),
    "check":             ("Expense",           "Payee Name"),
    "expense":           ("Expense",           "Payee Name"),
    "creditcardexpense": ("Expense",           "Payee Name"),
    "journalentry":      ("Manual Journal",    "Contact Name"),
    "deposit":           ("Manual Journal",    "Contact Name"),
    "creditcardcredit":  ("Manual Journal",    "Contact Name"),
    "salesreceipt":      ("Manual Journal",    "Customer Name"),
}

def _qbo_label(qbo_type):
    key = str(qbo_type or "bill").lower().replace(" ", "")
    return QBO_TYPE_LABELS.get(key, ("Transaction", "Contact Name"))

def _dominant_type_label(bills):
    """Pick the most common qbo_type label for naming the sheet."""
    from collections import Counter
    types = [str(b.get("qbo_type") or "bill").lower().replace(" ", "") for b in bills]
    if not types:
        return ("Transaction", "Contact Name")
    most = Counter(types).most_common(1)[0][0]
    return _qbo_label(most)


@app.route("/api/export/<realm_id>", methods=["GET"])
def api_export(realm_id: str):
    """
    Export as .xlsx with 2 sheets:
      Sheet 1 - QB Bills     : all bill fields + line items flattened (one row per line item)
      Sheet 2 - Attachments  : one row per file with download/upload status
    """
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from core.storage import load_index
    from flask import make_response

    idx = load_index(realm_id)
    if not idx.get("bills"):
        return jsonify({"error": "No data to export"}), 404

    bills   = list(idx["bills"].values())
    fetched = idx.get("fetched_at", "")

    # Type-aware labels for sheet + contact column
    type_label, contact_label = _dominant_type_label(bills)

    wb = Workbook()

    # ── Shared styles ──────────────────────────────────────────────────────────
    HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    HDR_FILL  = PatternFill("solid", start_color="1F4E79")
    BODY_FONT = Font(name="Arial", size=10)
    LEFT      = Alignment(horizontal="left", vertical="center")
    CENTER    = Alignment(horizontal="center", vertical="center")
    THIN      = Side(style="thin", color="D9D9D9")
    BORDER    = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    STATUS_FILL = {
        "matched":   PatternFill("solid", start_color="C6EFCE"),
        "no_match":  PatternFill("solid", start_color="FFCCCC"),
        "duplicate": PatternFill("solid", start_color="FFEB9C"),
        "manual":    PatternFill("solid", start_color="E2EFDA"),
        "pending":   PatternFill("solid", start_color="F2F2F2"),
        "success":   PatternFill("solid", start_color="C6EFCE"),
        "failed":    PatternFill("solid", start_color="FFCCCC"),
        "skipped":   PatternFill("solid", start_color="EEEEEE"),
    }

    def write_header(ws, headers):
        ws.append(headers)
        for cell in ws[1]:
            cell.font      = HDR_FONT
            cell.fill      = HDR_FILL
            cell.alignment = CENTER
            cell.border    = BORDER
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_num, status_col=None, status_val=None):
        for cell in ws[row_num]:
            cell.font      = BODY_FONT
            cell.border    = BORDER
            cell.alignment = LEFT
        if status_col and status_val and status_val in STATUS_FILL:
            ws.cell(row_num, status_col).fill = STATUS_FILL[status_val]

    def set_widths(ws, widths):
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — QB Bills (all fields, one row per line item)
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    _plural = {"Manual Journal":"Manual Journals","Expense":"Expenses",
               "Credit Note":"Credit Notes","Debit Note":"Debit Notes",
               "Invoice":"Invoices","Bill":"Bills","Transaction":"Transactions"}
    ws1.title = _plural.get(type_label, type_label + "s")
    ws1.freeze_panes = "A2"

    HDR1 = [
        f"QB {type_label} ID", "Doc Number",
        contact_label, "Contact ID",
        f"{type_label} Date", "Due Date", "Total Amount", "Balance",
        "Currency", "Exchange Rate",
        "AP Account", "Payment Terms", "Department",
        "Global Tax", "Total Tax",
        "Line #", "Line Description", "Line Amount",
        "GL Account", "Tax Code", "Billable", "Customer (line)",
        "Private Note", "Created At (QB)", "Last Updated (QB)",
        "Match Status", f"Wafeq {type_label} ID", "Match Note",
        "Attachment Count", "Fetched At",
    ]
    write_header(ws1, HDR1)

    MATCH_COL1 = HDR1.index("Match Status") + 1

    for b in bills:
        ms       = b.get("match_status", "pending")
        lines    = b.get("line_items", [])
        base     = [
            b.get("qb_bill_id", ""),
            b.get("doc_number", ""),
            b.get("vendor_name", ""),
            b.get("vendor_id", ""),
            b.get("txn_date", ""),
            b.get("due_date", ""),
            b.get("total_amt", ""),
            b.get("balance", ""),
            b.get("currency", ""),
            b.get("exchange_rate", 1),
            b.get("ap_account", ""),
            b.get("payment_terms", ""),
            b.get("department", ""),
            b.get("global_tax", ""),
            b.get("total_tax", ""),
        ]
        tail = [
            b.get("private_note", ""),
            b.get("created_at", ""),
            b.get("updated_at", ""),
            ms,
            b.get("wafeq_bill_id", ""),
            b.get("match_note", ""),
            len(b.get("attachments", [])),
            fetched,
        ]

        if not lines:
            # No line items — write single row with blank line columns
            row = base + ["", "", "", "", "", "", ""] + tail
            ws1.append(row)
            style_row(ws1, ws1.max_row, MATCH_COL1, ms)
        else:
            for ln in lines:
                row = base + [
                    ln.get("line_num", ""),
                    ln.get("description", ""),
                    ln.get("amount", ""),
                    ln.get("account", ""),
                    ln.get("tax_code", ""),
                    ln.get("billable", ""),
                    ln.get("customer", ""),
                ] + tail
                ws1.append(row)
                style_row(ws1, ws1.max_row, MATCH_COL1, ms)

    set_widths(ws1, [
        16, 14,          # Bill ID, Doc Number
        26, 14,          # Vendor name, ID
        12, 12,          # Dates
        14, 12, 10, 12,  # Amounts, currency, FX
        20, 16, 16,      # AP, terms, dept
        16, 10,          # Tax
        8, 30, 12,       # Line #, desc, amount
        24, 10, 12, 20,  # GL, tax code, billable, customer
        30, 20, 20,      # Note, created, updated
        14, 30, 36,      # Match, wafeq id, note
        14, 22,          # Att count, fetched at
    ])

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Attachments (one row per file)
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Attachments")
    ws2.freeze_panes = "A2"

    HDR2 = [
        f"QB {type_label} ID", "Doc Number", contact_label,
        f"{type_label} Date", "Total Amount", "Currency",
        "Match Status", f"Wafeq {type_label} ID",
        # File
        "Attachable ID", "File Name", "Content Type",
        "File Size (KB)", "Note", "Include On Send",
        "Created At (QB)",
        # Migration
        "Download Status", "Upload Status",
        "Wafeq File ID", "Upload Error",
        "Fetched At",
    ]
    write_header(ws2, HDR2)

    MATCH_COL2  = HDR2.index("Match Status")  + 1
    UPLOAD_COL2 = HDR2.index("Upload Status") + 1

    for b in bills:
        ms   = b.get("match_status", "pending")
        atts = b.get("attachments", [])

        if not atts:
            row = [
                b.get("qb_bill_id", ""), b.get("doc_number", ""),
                b.get("vendor_name", ""), b.get("txn_date", ""),
                b.get("total_amt", ""), b.get("currency", ""),
                ms, b.get("wafeq_bill_id", ""),
                "", "(no attachments)", "", "", "", "", "",
                "", "", "", "", fetched,
            ]
            ws2.append(row)
            style_row(ws2, ws2.max_row, MATCH_COL2, ms)
        else:
            for att in atts:
                us  = att.get("upload_status", "pending")
                sz  = att.get("file_size_bytes") or 0
                row = [
                    b.get("qb_bill_id", ""),
                    b.get("doc_number", ""),
                    b.get("vendor_name", ""),
                    b.get("txn_date", ""),
                    b.get("total_amt", ""),
                    b.get("currency", ""),
                    ms,
                    b.get("wafeq_bill_id", ""),
                    att.get("attachable_id", ""),
                    att.get("file_name", ""),
                    att.get("content_type", ""),
                    round(sz / 1024, 1),
                    att.get("note", ""),
                    att.get("include_on_send", ""),
                    att.get("created_at", ""),
                    att.get("download_status", ""),
                    us,
                    att.get("wafeq_file_id", ""),
                    att.get("error", ""),
                    fetched,
                ]
                ws2.append(row)
                rn = ws2.max_row
                style_row(ws2, rn, MATCH_COL2, ms)
                if us in STATUS_FILL:
                    ws2.cell(rn, UPLOAD_COL2).fill = STATUS_FILL[us]

    set_widths(ws2, [
        16, 14, 26,      # Bill ID, doc, vendor
        12, 14, 10,      # date, amt, currency
        14, 30,          # match status, wafeq id
        20, 30, 22,      # attachable id, filename, content type
        12, 20, 14, 20,  # size, note, include, created
        16, 14,          # download, upload status
        30, 36, 22,      # wafeq file id, error, fetched at
    ])

    # ── Stream response ────────────────────────────────────────────────────────
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    resp = make_response(output.read())
    resp.headers["Content-Type"] =         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    resp.headers["Content-Disposition"] =         f"attachment; filename=migration_report_{realm_id[:8]}.xlsx"
    return resp
# ---------------------------------------------------------------------------
# Simple session-based auth
# ---------------------------------------------------------------------------




if __name__ == "__main__":
    print("\n" + "=" * 55)
    print("  QBO → Wafeq  |  http://localhost:" + str(FLASK_PORT))
    print("=" * 55)
    print(f"  Callback : {REDIRECT_URI}")
    print(f'  ngrok    : ngrok http {FLASK_PORT} --request-header-add "ngrok-skip-browser-warning:true"')
    print("=" * 55 + "\n")
    app.run(host="0.0.0.0", port=FLASK_PORT, debug=False, use_reloader=False)